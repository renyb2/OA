#    Copyright 2022 Renyb && Liyr
#
#    Licensed under the Apache License, Version 2.0 (the "License"); you may
#    not use this file except in compliance with the License. You may obtain
#    a copy of the License at
#
#         http://www.apache.org/licenses/LICENSE-2.0
#
#    Unless required by applicable law or agreed to in writing, software
#    distributed under the License is distributed on an "AS IS" BASIS, WITHOUT
#    WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
#    License for the specific language governing permissions and limitations
#    under the License.

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment


class Excel(object):
    """ Excel Manager """

    def __init__(self, file, sheet=None) -> None:
        self.file = file
        self.sheet = sheet
        
        self._load()

    def _load(self) -> None:
        """ load excel sheet
        """
        self.wb = load_workbook(self.file, data_only=True)
        self.ws = self.wb[self.sheet] if self.sheet else self.wb.active

    def _save(self) -> bool:
        """ save excel

        Returns:
            bool: save success or not
        """
        return self.wb.save(self.file)

    def _format_style(self, cell) -> None:
        border_type = Side(
            border_style="thin", 
            color="000000"
        )
        border = Border(
            left=border_type,
            right=border_type,
            top=border_type,
            bottom=border_type
        )
        font = Font(
            size=12, 
            bold=False, 
            name='微软雅黑', 
            color="000000"
        )
        alignment = Alignment(
            horizontal='center', 
            vertical='center', 
            wrap_text=True
        )

        cell.border = border
        cell.font = font
        cell.alignment = alignment

    def get_max_row(self) -> int:
        """ get all row counts in excel

        Returns:
            int: row counts
        """
        return self.ws.max_row

    def add_row_with_values(self, value, format=False) -> bool:
        """ add a row of data to excel

        Args:
            value (list): all values of a row

        Returns:
            bool: add success or not
        """
        self.ws.append(value)
        if format:
            row = self.get_max_row()
            for cell in self.ws[row]:
                self._format_style(cell)
        return self._save()

    def get_row_index_by_value(self, range) -> dict:
        """ Generate a map, specifying the range of values and their indices

        Args:
            range (str): cell ranges, split by ':'. eg: 'E13:E57'.

        Returns:
            dict: a map of values and their indices
        """
        record = {}
        for cell in self.ws[range]:
            value = cell[0].value
            row = cell[0].row
            if value:
                if value in record.keys():
                    record[value].append(row)
                else:
                    record[value] = [row]
        return record

    def get_row_values_by_index(self, row) -> list:
        """ get all the data of the specified row

        Args:
            row (int): line number

        Returns:
            list: all values of a row
        """
        return [item.value for item in self.ws[row]]

    def delete_row_by_index(self, row) -> bool:
        """ delete the specified row

        Args:
            row (int): line number

        Returns:
            bool: delete success or not
        """
        self.ws.delete_rows(row)
        return self._save()
